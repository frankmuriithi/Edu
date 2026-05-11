from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import get_user_model, authenticate, login, logout
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.utils.timezone import now, localtime
from django.contrib import messages
from datetime import date
from .models import Course, ClassSession, Attendance, Notification, Profile, Message
from .forms import (
    ManualCheckinForm, RegistrationForm, ClassSessionForm, MessageForm,
    NotificationForm, UserUpdateForm, StudentCourseEnrollmentForm, ManageCourseStudentsForm, CourseForm, CustomLoginForm
)
from django.contrib.auth.forms import AuthenticationForm
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from django.db.models import Q
from django.contrib.auth.models import User

from django.utils.timezone import make_aware
from datetime import datetime

# -------------------------
# STATUS WEIGHTS 🔥
# -------------------------
ATTENDANCE_WEIGHTS = {
    "present": 1,
    "late": 0.5,
    "absent": 0
}
ATTENDED_STATUSES = ['present', 'late']

def get_session_status(session):
    now_dt = localtime(now())

    start_dt = make_aware(datetime.combine(session.date, session.start_time))
    end_dt = make_aware(datetime.combine(session.date, session.end_time))

    if now_dt < start_dt:
        return "UPCOMING"
    elif now_dt > end_dt:
        return "CLOSED"
    else:
        return "ACTIVE"

# -------------------------
# DASHBOARD
# -------------------------
@login_required
def dashboard(request):
    today = localtime(now()).date()
    role = request.user.profile.role

    if role == 'LECTURER':
        today_sessions = ClassSession.objects.filter(course__lecturer=request.user, date=today)
    else:
        today_sessions = ClassSession.objects.filter(course__students=request.user, date=today)

    session_data = []
    unchecked_sessions = []

    for session in today_sessions:
        status = get_session_status(session)

        attendance = None
        if role == 'STUDENT':
            attendance = Attendance.objects.filter(student=request.user, session=session).first()

            if status in ["ACTIVE", "CLOSED"] and not attendance:
                unchecked_sessions.append(session)

        session_data.append({
            "session": session,
            "attendance": attendance,
            "session_status": status
        })

    # -------------------------
    # STUDENT STATS (WEIGHTED)
    # -------------------------
    overall_attendance = 0

    if role == "STUDENT":
        records = Attendance.objects.filter(student=request.user)

        total = records.count()
        weighted_sum = sum(ATTENDANCE_WEIGHTS.get(r.status, 0) for r in records)
        
        overall_attendance = round((weighted_sum / total) * 100) if total else 0


    # -------------------------
    # LECTURER STATS
    # -------------------------
    lecturer_stats = None

    if role == "LECTURER":
        active_session = today_sessions.filter(
            date=today
        ).first()

        if active_session:
            total_students = active_session.course.students.count()

            checked = Attendance.objects.filter(
                session=active_session,
                status__in=ATTENDED_STATUSES
            ).count()

            absent = total_students - checked

            percentage = (checked / total_students) * 100 if total_students else 0

            lecturer_stats = {
                "attendance_percentage": round(percentage, 2),
                "checked_in_students": checked,
                "total_students": total_students,
                "absent_students": absent
            }

    return render(request, 'dashboard/dashboard.html', {
        "session_data": session_data,
        "unchecked_sessions": unchecked_sessions,
        "overall_attendance": overall_attendance,
        "lecturer_stats": lecturer_stats,
        "role": role
    })


# -------------------------
# Attendance List
# -------------------------
@login_required
def attendance_list(request):
    role = request.user.profile.role

    if role == "STUDENT":
        courses = request.user.courses_enrolled.all()
        attendances = Attendance.objects.filter(student=request.user, course__in=courses)

    elif role == "LECTURER":
        courses = Course.objects.filter(lecturer=request.user)
        attendances = Attendance.objects.filter(course__in=courses)

    else:
        courses = Course.objects.all()
        attendances = Attendance.objects.all()

    selected_course = request.GET.get("course")
    if selected_course:
        attendances = attendances.filter(course_id=selected_course)

    return render(request, "dashboard/attendance_list.html", {
        "courses": courses,
        "attendances": attendances,
        "selected_course": selected_course
    })

# -------------------------
# Search
# -------------------------
@login_required
def search_results(request):
    query = request.GET.get('q', '').strip()

    users = User.objects.none()
    profiles = Profile.objects.none()
    courses = Course.objects.none()
    sessions = ClassSession.objects.none()
    attendances = Attendance.objects.none()
    messages_qs = Message.objects.none()

    if query:

        # -------------------------
        # USERS (students, lecturers, admin if any)
        # -------------------------
        users = User.objects.filter(
            Q(username__icontains=query) |
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(email__icontains=query)
        )

        # -------------------------
        # PROFILES
        # -------------------------
        profiles = Profile.objects.filter(
            Q(role__icontains=query) |
            Q(registration_number__icontains=query) |
            Q(staff_number__icontains=query) |
            Q(location__icontains=query) |
            Q(bio__icontains=query)
        )

        # -------------------------
        # COURSES
        # -------------------------
        courses = Course.objects.filter(
            Q(name__icontains=query) |
            Q(code__icontains=query) |
            Q(description__icontains=query) |
            Q(lecturer__username__icontains=query)
        )

        # -------------------------
        # CLASS SESSIONS (FIXED — NO title field!)
        # -------------------------
        sessions = ClassSession.objects.filter(
            Q(course__name__icontains=query) |
            Q(course__code__icontains=query) |
            Q(location__icontains=query) |
            Q(date__icontains=query)
        )

        # -------------------------
        # ATTENDANCE
        # -------------------------
        attendances = Attendance.objects.filter(
            Q(student__username__icontains=query) |
            Q(course__name__icontains=query) |
            Q(status__icontains=query) |
            Q(date__icontains=query)
        )

        # -------------------------
        # MESSAGES
        # -------------------------
        messages_qs = Message.objects.filter(
            Q(content__icontains=query) |
            Q(sender__username__icontains=query) |
            Q(recipient__username__icontains=query)
        )

    return render(request, 'dashboard/search_results.html', {
        "query": query,
        "users": users,
        "profiles": profiles,
        "courses": courses,
        "sessions": sessions,
        "attendances": attendances,
        "messages_qs": messages_qs,
    })
# -------------------------
# Teacher: Students Checked In
# -------------------------
@login_required
def teacher_students_checked_in(request):
    if request.user.profile.role != 'LECTURER':
        return HttpResponse("Unauthorized", status=403)

    courses = request.user.courses_taught.all()
    sessions = ClassSession.objects.filter(course__in=courses)

    student_ids = Attendance.objects.filter(
        session__in=sessions,
        status__in=ATTENDED_STATUSES
    ).values_list('student_id', flat=True).distinct()

    unique_students = User.objects.filter(id__in=student_ids)

    return render(request, 'dashboard/teacher_students_checked_in.html', {
        'checked_in_students': unique_students
    })

# -------------------------
# Check-in (via UI button)
# -------------------------
@login_required
def checkin(request, session_id):
    if request.method != "POST":
        return JsonResponse({"success": False}, status=405)

    session = get_object_or_404(ClassSession, id=session_id)

    if get_session_status(session) != "ACTIVE":
        messages.error(request, "This session is not active.")
        return redirect("dashboard")

    current_dt = localtime(now())

    Attendance.objects.update_or_create(
        student=request.user,
        session=session,
        defaults={
            "course": session.course,
            "date": session.date,
            "time": current_dt.time(),
            "status": "present",
            "marked_by": request.user
        }
    )

    return redirect("checkin_success")

@login_required
def checkin_success(request):
    return render(request, 'dashboard/checkin_success.html')    
# -------------------------
# MANUAL CHECK-IN (FIXED 🔥)
# -------------------------
@login_required
def manual_checkin(request):
    if request.user.profile.role != "LECTURER":
        return redirect("dashboard")

    if request.method == "POST":
        form = ManualCheckinForm(request.POST, user=request.user)

        if form.is_valid():
            student = form.cleaned_data["student"]
            course = form.cleaned_data["course"]
            date_val = form.cleaned_data["date"]
            time_val = form.cleaned_data["time"]
            status = form.cleaned_data["status"]
            reason = form.cleaned_data["reason"]

            session = ClassSession.objects.filter(
                course=course,
                date=date_val
            ).order_by("-start_time").first()

            if not session:
                messages.error(request, "No session found.")
                return redirect("dashboard")

            # 🔥 IMPORTANT FIX: allow editing even if CLOSED
            Attendance.objects.update_or_create(
                student=student,
                session=session,
                defaults={
                    "course": course,
                    "date": date_val,
                    "time": time_val,
                    "status": status,
                    "reason": reason,
                    "marked_by": request.user
                }
            )

            messages.success(request, "Attendance updated successfully.")
            return redirect("dashboard")

    else:
        form = ManualCheckinForm(user=request.user)

    return render(request, "dashboard/manual_checkin_modal.html", {"form": form})

# -------------------------
# Create Class Session (Lecturer)
# -------------------------
@login_required
def create_class_session(request):
    if request.user.profile.role != "LECTURER":
        return redirect("dashboard")

    if request.method == "POST":
        form = ClassSessionForm(request.user, request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Session created")
            return redirect("dashboard")
    else:
        form = ClassSessionForm(request.user)

    return render(request, "dashboard/create_class_session.html", {"form": form})

# -------------------------
# Edit / Delete Class Session (Lecturer)
# -------------------------
def is_teacher(user):
    return user.profile.role == 'LECTURER'


@user_passes_test(is_teacher)
@login_required
def edit_class_session(request, session_id):
    session = get_object_or_404(ClassSession, pk=session_id)

    if request.method == 'POST':
        form = ClassSessionForm(request.user, request.POST, instance=session)
        if form.is_valid():
            form.save()
            messages.success(request, "Class session updated successfully.")
            return redirect('dashboard')
    else:
        form = ClassSessionForm(request.user, instance=session)

    return render(request, 'dashboard/edit_class_session.html', {'form': form, 'session': session})


@user_passes_test(is_teacher)
@login_required
def delete_class_session(request, session_id):
    session = get_object_or_404(ClassSession, pk=session_id)

    if request.method == 'POST':
        session.delete()
        messages.success(request, "Class session deleted successfully.")
        return redirect('dashboard')

    return render(request, 'dashboard/confirm_delete_session.html', {'session': session})


@login_required
def courses(request):
    role = request.user.profile.role

    # -------------------------
    # STUDENT LOGIC
    # -------------------------
    if role == 'STUDENT':
        enrolled_courses = request.user.courses_enrolled.all()
        available_courses = Course.objects.exclude(students=request.user)

        if request.method == 'POST':
            action = request.POST.get('action')
            course_id = request.POST.get('course_id')

            if course_id:
                course = get_object_or_404(Course, id=course_id)

                if action == 'enroll':
                    course.students.add(request.user)
                    messages.success(request, f"You have enrolled in {course.name}")

                elif action == 'unenroll':
                    course.students.remove(request.user)
                    messages.success(request, f"You have unenrolled from {course.name}")

            return redirect('courses')

        context = {
            'role': role,
            'enrolled_courses': enrolled_courses,
            'available_courses': available_courses,
        }

    # -------------------------
    # LECTURER LOGIC
    # -------------------------
    elif role == 'LECTURER':
        lecturer_courses = Course.objects.filter(lecturer=request.user)
        all_students = User.objects.filter(profile__role='STUDENT')

        if request.method == 'POST':
            action = request.POST.get('action')
            course_id = request.POST.get('course_id')
            student_id = request.POST.get('student_id')

            if course_id:
                course = get_object_or_404(Course, id=course_id, lecturer=request.user)

                if action == 'add_student' and student_id:
                    student = get_object_or_404(User, id=student_id)
                    course.students.add(student)
                    messages.success(request, f"Student added to {course.name}")

                elif action == 'remove_student' and student_id:
                    student = get_object_or_404(User, id=student_id)
                    course.students.remove(student)
                    messages.success(request, f"Student removed from {course.name}")

            return redirect('courses')

        context = {
            'role': role,
            'lecturer_courses': lecturer_courses,
            'all_students': all_students,
        }

    # -------------------------
    # ADMIN LOGIC
    # -------------------------
    else:
        all_courses = Course.objects.all()

        context = {
            'role': role,
            'all_courses': all_courses,
        }

    return render(request, 'dashboard/courses.html', context)

# -------------------------
# Add Course (Lecturer)
# -------------------------
@login_required
def add_course(request):
    if request.user.profile.role != 'LECTURER':
        return HttpResponse("Unauthorized", status=403)

    if request.method == 'POST':
        form = CourseForm(request.POST)

        if form.is_valid():
            course = form.save(commit=False)
            course.lecturer = request.user  # auto assign lecturer
            course.save()
            form.save_m2m()

            print("COURSE SAVED:", course)  # debug
            messages.success(request, "Course created successfully!")
            return redirect('courses')
        else:
            print("FORM ERRORS:", form.errors)
            messages.error(request, form.errors)

    else:
        form = CourseForm()

    return render(request, 'dashboard/add_course.html', {'form': form})
# -------------------------
# Notifications
# -------------------------
@login_required
def send_notification(request):
    if request.user.profile.role == 'STUDENT':
        return redirect('view_notifications')

    if request.method == 'POST':
        form = NotificationForm(request.POST, sender=request.user)
        if form.is_valid():
            notif = form.save(commit=False)
            notif.sender = request.user
            notif.save()
            return redirect('view_notifications')
    else:
        form = NotificationForm(sender=request.user)

    return render(request, 'dashboard/send_notification.html', {'form': form})


@login_required
def view_notifications(request):
    notifications = Notification.objects.filter(
        recipient=request.user
    ).order_by('-timestamp')

    # ✅ MARK ALL AS READ WHEN PAGE IS OPENED
    notifications.update(is_read=True)

    return render(request, 'dashboard/view_notifications.html', {
        'notifications': notifications
    })

@require_POST
@login_required
def mark_notification_read(request, notification_id):
    notification = get_object_or_404(Notification, pk=notification_id, recipient=request.user)
    notification.is_read = True
    notification.save()
    return JsonResponse({'success': True})


# -------------------------
# REPORTS (WEIGHTED 🔥)
# -------------------------
@login_required
def reports(request):
    role = request.user.profile.role
    attendance_data = []

    if role == 'STUDENT':
        courses = request.user.courses_enrolled.all()

        for course in courses:
            records = Attendance.objects.filter(
                student=request.user,
                session__course=course
            )

            total = records.count()
            weighted = sum(ATTENDANCE_WEIGHTS.get(r.status, 0) for r in records)

            percentage = round((weighted / total) * 100) if total else 0

            attendance_data.append({
                'course': course.name,
                'attended': weighted,
                'total': total,
                'percentage': percentage
            })

        return render(request, 'dashboard/reports.html', {'attendance_data': attendance_data})

    elif role == 'LECTURER':
        courses = Course.objects.filter(lecturer=request.user)

        for course in courses:
            sessions = ClassSession.objects.filter(course=course)

            for session in sessions:
                count = Attendance.objects.filter(
                    session=session,
                    status__in=ATTENDED_STATUSES
                ).count()

                attendance_data.append({
                    'course': course.name,
                    'session_date': session.date,
                    'start_time': session.start_time,
                    'end_time': session.end_time,
                    'students_present': count
                })

        return render(request, 'dashboard/report_teacher.html', {'attendance_data': attendance_data})


# -------------------------
# EXPORT REPORT (FIXED)
# -------------------------
@login_required
def export_reports_excel(request):
    role = request.user.profile.role

    if role == 'STUDENT':
        courses = request.user.courses_enrolled.all()
    else:
        courses = Course.objects.filter(lecturer=request.user)

    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(["Course", "Score", "Total", "Attendance %"])

    for course in courses:
        records = Attendance.objects.filter(
            student=request.user,
            course=course
        )

        total = records.count()
        weighted = sum(ATTENDANCE_WEIGHTS.get(r.status, 0) for r in records)

        percent = round((weighted / total) * 100) if total else 0

        ws.append([course.name, weighted, total, percent])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=attendance.xlsx'
    wb.save(response)
    return response
# -------------------------
# Export Teacher Report (fixed: was using Course.objects.filter(teacher=...) — now 'lecturer')
# -------------------------
@login_required
def export_teacher_report_excel(request):
    if request.user.profile.role != 'LECTURER':
        return HttpResponse("Unauthorized", status=403)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Teacher Report"

    headers = ['Course', 'Date', 'Start Time', 'End Time', 'Students Present']
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    # Fixed: changed 'teacher' to 'lecturer' to match the Course model field
    courses = Course.objects.filter(lecturer=request.user)
    for course in courses:
        sessions = ClassSession.objects.filter(course=course).order_by('-date')
        for session in sessions:
            students_present = Attendance.objects.filter(session=session, status='present').count()
            ws.append([
                course.name,
                session.date.strftime('%Y-%m-%d'),
                str(session.start_time),
                str(session.end_time),
                students_present
            ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=teacher_report.xlsx'
    wb.save(response)
    return response


# -------------------------
# Export Attendance History (all roles)
# -------------------------
@login_required
def export_attendance_excel(request):
    role = request.user.profile.role
    if role == 'STUDENT':
        attendances = Attendance.objects.filter(student=request.user)
    elif role == 'LECTURER':
        attendances = Attendance.objects.filter(course__lecturer=request.user)
    else:
        attendances = Attendance.objects.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance History"

    ws.append(['Student', 'Course', 'Date', 'Time', 'Status', 'Reason'])

    for att in attendances:
        ws.append([
            att.student.username,
            att.course.name,
            att.date.strftime('%Y-%m-%d'),
            att.time.strftime('%H:%M'),
            att.status,
            att.reason or ''
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=attendance_history.xlsx'
    wb.save(response)
    return response


# -------------------------
# Messages
# -------------------------
@login_required
def messages_page(request):
    return render(request, 'dashboard/messages.html')


@login_required
def message_view(request):
    user_messages = Message.objects.filter(
        recipient=request.user
    ).order_by('-timestamp')

    # ✅ MARK AS READ
    user_messages.update(is_read=True)

    form = MessageForm()

    if request.method == 'POST':
        form = MessageForm(request.POST)
        if form.is_valid():
            msg = form.save(commit=False)
            msg.sender = request.user
            msg.save()
            return redirect('messages')

    return render(request, 'dashboard/message.html', {
        'messages': user_messages,
        'form': form,
    })


# -------------------------
# Settings
# -------------------------
@login_required
def settings(request):
    if request.method == 'POST':
        form = UserUpdateForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Your account was updated successfully!')
            return redirect('settings')
    else:
        form = UserUpdateForm(instance=request.user)

    return render(request, 'dashboard/settings.html', {'form': form})


# -------------------------
# Auth: Register, Login, Logout
# -------------------------
def register(request):
    if request.method == 'POST':
        form = RegistrationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Registration successful. You can now log in.")
            return redirect('login')
    else:
        form = RegistrationForm()
    return render(request, 'dashboard/register.html', {'form': form})





def user_login(request):
    if request.method == "POST":
        email = request.POST.get('email')
        password = request.POST.get('password')

        try:
            user_obj = User.objects.get(email=email)
            user = authenticate(request, username=user_obj.username, password=password)

            if user is not None:
                login(request, user)
                return redirect('dashboard')
            else:
                messages.error(request, "Invalid password")

        except User.DoesNotExist:
            messages.error(request, "User with this email does not exist")

    return render(request, 'dashboard/login.html')

def user_logout(request):
    logout(request)
    return redirect('login')